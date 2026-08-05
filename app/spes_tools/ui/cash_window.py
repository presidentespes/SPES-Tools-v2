from __future__ import annotations

import csv
from pathlib import Path

from PySide6.QtWidgets import (
    QComboBox,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from spes_tools.services.storage import add_history, load_abi_config


HEADERS = [
    "DATA",
    "NUMERO",
    "DESCRIZIONE",
    "ENTRATA",
    "USCITA",
    "CAUSALE",
    "CAUSALE ABI",
    "SOGGETTO",
    "NOTE",
]
ABI_COLUMN = HEADERS.index("CAUSALE ABI")
OLD_HEADERS = ["DATA", "NUMERO", "DESCRIZIONE", "ENTRATA", "USCITA", "CAUSALE", "SOGGETTO", "NOTE"]


class CashWindow(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("SPES Configuratore Contabile - Gestione cassa")
        self.resize(1240, 720)
        self._build_ui()
        self.add_row()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        title = QLabel("Gestione cassa")
        title.setStyleSheet("font-size: 27px; font-weight: bold; color: #073b84;")
        layout.addWidget(title)

        subtitle = QLabel(
            "Registro entrate e uscite con causali ABI CASSA separate da NEXI, BCC e VOLKSBANK"
        )
        subtitle.setStyleSheet("color: #53657a;")
        layout.addWidget(subtitle)

        buttons = QHBoxLayout()
        for text, callback in (
            ("Nuova riga", self.add_row),
            ("Elimina riga", self.delete_selected_rows),
            ("Importa", self.import_file),
            ("Esporta CSV", self.export_csv),
            ("Esporta Excel", self.export_xlsx),
            ("Azzera", self.clear_rows),
        ):
            button = QPushButton(text)
            button.clicked.connect(callback)
            buttons.addWidget(button)
        buttons.addStretch()
        layout.addLayout(buttons)

        self.table = QTableWidget(0, len(HEADERS))
        self.table.setHorizontalHeaderLabels(HEADERS)
        self.table.setAlternatingRowColors(True)
        self.table.itemChanged.connect(self.update_totals)
        self.table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.table)

        self.totals = QLabel()
        self.totals.setStyleSheet(
            "font-size: 16px; font-weight: 650; color: #073b84; "
            "background: #f3f6f9; padding: 10px;"
        )
        layout.addWidget(self.totals)
        self.update_totals()

    def _cash_codes(self) -> list[str]:
        return list(load_abi_config().get("CASSA", {}).values())

    def _abi_combo(self, selected: str = "") -> QComboBox:
        combo = QComboBox()
        combo.setEditable(False)
        combo.addItem("")
        combo.addItems(self._cash_codes())
        if selected:
            index = combo.findText(selected)
            if index < 0:
                combo.addItem(selected)
                index = combo.findText(selected)
            combo.setCurrentIndex(index)
        combo.currentTextChanged.connect(self.update_totals)
        return combo

    def add_row(self, values: list[str] | None = None) -> None:
        row = self.table.rowCount()
        self.table.insertRow(row)
        values = (values or [""] * len(HEADERS))[: len(HEADERS)]
        values += [""] * (len(HEADERS) - len(values))
        for col, value in enumerate(values):
            if col == ABI_COLUMN:
                self.table.setCellWidget(row, col, self._abi_combo(str(value)))
            else:
                self.table.setItem(row, col, QTableWidgetItem(str(value)))
        if not values[1]:
            self.table.item(row, 1).setText(str(row + 1))
        self.update_totals()

    def delete_selected_rows(self) -> None:
        rows = sorted({index.row() for index in self.table.selectedIndexes()}, reverse=True)
        for row in rows:
            self.table.removeRow(row)
        self.update_totals()

    def clear_rows(self) -> None:
        self.table.setRowCount(0)
        self.add_row()

    def import_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Importa registro cassa",
            "",
            "Registro cassa (*.csv *.xlsx);;CSV (*.csv);;Excel (*.xlsx)",
        )
        if not path:
            return
        try:
            rows = self._read_xlsx(Path(path)) if path.lower().endswith(".xlsx") else self._read_csv(Path(path))
        except Exception as exc:
            QMessageBox.critical(self, "Errore importazione", str(exc))
            return
        self.table.setRowCount(0)
        for row in rows:
            self.add_row(row)
        if not rows:
            self.add_row()
        self.update_totals()

    def export_csv(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Esporta registro cassa", "Registro_Cassa_SPES_con_ABI.csv", "CSV (*.csv)"
        )
        if not path:
            return
        if not path.lower().endswith(".csv"):
            path += ".csv"
        try:
            with Path(path).open("w", newline="", encoding="utf-8-sig") as handle:
                writer = csv.writer(handle, delimiter=";", lineterminator="\n")
                writer.writerow(HEADERS)
                writer.writerows(self._rows())
            self._record_history(path)
        except Exception as exc:
            QMessageBox.critical(self, "Errore esportazione", str(exc))
            return
        QMessageBox.information(self, "Esportazione completata", path)

    def export_xlsx(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Esporta registro cassa", "Registro_Cassa_SPES_con_ABI.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Registro Cassa"
            ws.append(HEADERS)
            for row in self._rows():
                ws.append(row)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="17365D")

            widths = [13, 10, 36, 14, 14, 25, 16, 28, 36]
            for index, width in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(index)].width = width

            codes = self._cash_codes()
            lookup = wb.create_sheet("Causali Cassa")
            lookup.append(["CAUSALE ABI", "TIPO MOVIMENTO"])
            for key, code in load_abi_config().get("CASSA", {}).items():
                lookup.append([code, key.replace("_", " ").title()])
            formula = '"' + ",".join(codes) + '"'
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(validation)
            validation.add(f"G2:G{max(500, ws.max_row + 100)}")
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:I{max(2, ws.max_row)}"
            wb.save(path)
            self._record_history(path)
        except Exception as exc:
            QMessageBox.critical(self, "Errore esportazione", str(exc))
            return
        QMessageBox.information(self, "Esportazione completata", path)

    def update_totals(self, *_args) -> None:
        total_in = sum(self._parse_amount(self._cell(row, 3)) for row in range(self.table.rowCount()))
        total_out = sum(self._parse_amount(self._cell(row, 4)) for row in range(self.table.rowCount()))
        balance = total_in - total_out
        self.totals.setText(
            f"Totale entrate: EUR {total_in:,.2f}    |    "
            f"Totale uscite: EUR {total_out:,.2f}    |    "
            f"Saldo cassa: EUR {balance:,.2f}"
        )

    def _rows(self) -> list[list[str]]:
        result: list[list[str]] = []
        for row in range(self.table.rowCount()):
            values = [self._cell(row, col).strip() for col in range(self.table.columnCount())]
            if any(values):
                result.append(values)
        return result

    def _cell(self, row: int, col: int) -> str:
        if col == ABI_COLUMN:
            widget = self.table.cellWidget(row, col)
            return widget.currentText() if isinstance(widget, QComboBox) else ""
        item = self.table.item(row, col)
        return item.text() if item else ""

    @staticmethod
    def _parse_amount(value: str) -> float:
        value = value.strip().replace("EUR", "").replace("€", "").replace(" ", "")
        if not value:
            return 0.0
        if "," in value:
            value = value.replace(".", "").replace(",", ".")
        try:
            return float(value)
        except ValueError:
            return 0.0

    @staticmethod
    def _normalize_import_rows(rows: list[list[str]]) -> list[list[str]]:
        result: list[list[str]] = []
        for row in rows:
            if len(row) == len(OLD_HEADERS):
                row = row[:6] + [""] + row[6:]
            result.append((row + [""] * len(HEADERS))[: len(HEADERS)])
        return result

    @classmethod
    def _read_csv(cls, path: Path) -> list[list[str]]:
        text = path.read_text(encoding="utf-8-sig", errors="ignore")
        first = text.splitlines()[0] if text.splitlines() else ""
        delimiter = ";" if ";" in first else ","
        rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
        if rows:
            header = [x.strip().upper() for x in rows[0]]
            if header[: len(HEADERS)] == HEADERS or header[: len(OLD_HEADERS)] == OLD_HEADERS:
                rows = rows[1:]
        return cls._normalize_import_rows([row for row in rows if any(cell.strip() for cell in row)])

    @classmethod
    def _read_xlsx(cls, path: Path) -> list[list[str]]:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True)
        ws = wb["Registro Cassa"] if "Registro Cassa" in wb.sheetnames else wb.active
        raw = [["" if value is None else str(value) for value in row] for row in ws.iter_rows(values_only=True)]
        if raw:
            header = [x.strip().upper() for x in raw[0]]
            if header[: len(HEADERS)] == HEADERS or header[: len(OLD_HEADERS)] == OLD_HEADERS:
                raw = raw[1:]
        return cls._normalize_import_rows([row for row in raw if any(cell.strip() for cell in row)])

    def _record_history(self, path: str) -> None:
        add_history(
            module="Gestione cassa",
            source="Registro manuale",
            output=path,
            rows=len(self._rows()),
            details=self.totals.text(),
        )
