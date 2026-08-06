from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import unicodedata
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Iterable
from urllib.parse import urljoin, urlparse

import pdfplumber
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from spes_tools.services.storage import data_dir

FGI_CLASSIFICATIONS_URL = (
    "https://www.federginnastica.it/calendario-gare/classifiche/classifiche.html"
)
FGI_CLUB_CODE = "000112"
FGI_CLUB_NAMES = (
    "SPES MESTRE",
    "SPES MESTRE GINNASTICA",
    "SPES MESTRE GINNASTICA A.S.D.",
    "SPES MESTRE GINNASTICA ASD",
)

_MONTHS = {
    "gennaio": 1,
    "febbraio": 2,
    "marzo": 3,
    "aprile": 4,
    "maggio": 5,
    "giugno": 6,
    "luglio": 7,
    "agosto": 8,
    "settembre": 9,
    "ottobre": 10,
    "novembre": 11,
    "dicembre": 12,
}

_DISCIPLINES = (
    "Artistica Maschile",
    "Artistica Femminile",
    "Ritmica",
    "Trampolino",
    "Aerobica",
    "Ginnastica per tutti",
    "TeamGym",
    "Parkour",
    "Ginnastica Acrobatica",
)


@dataclass(frozen=True)
class FgiResult:
    result_id: str
    season: str
    date: str
    discipline: str
    competition: str
    category: str
    athlete: str
    club: str
    position: str
    score: str
    apparatus: str
    source_url: str
    document_name: str
    raw_row: str
    updated_at: str


@dataclass(frozen=True)
class UpdateReport:
    scanned_documents: int
    matching_rows: int
    added_results: int
    total_results: int
    warnings: tuple[str, ...] = ()


def results_path() -> Path:
    return data_dir() / "fgi_results.json"


def current_season(today: date | None = None) -> tuple[date, date, str]:
    today = today or date.today()
    start_year = today.year if today.month >= 9 else today.year - 1
    start = date(start_year, 9, 1)
    end = date(start_year + 1, 8, 31)
    return start, end, f"{start_year}/{start_year + 1}"


def load_results() -> list[FgiResult]:
    path = results_path()
    if not path.exists():
        return []
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    results: list[FgiResult] = []
    for item in raw if isinstance(raw, list) else []:
        if not isinstance(item, dict):
            continue
        try:
            results.append(FgiResult(**item))
        except TypeError:
            continue
    return results


def save_results(results: Iterable[FgiResult]) -> None:
    payload = [asdict(item) for item in results]
    results_path().write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def update_fgi_results(
    start_date: date,
    end_date: date,
    *,
    progress: Callable[[str], None] | None = None,
    session: requests.Session | None = None,
) -> UpdateReport:
    """Download FGI classifications and keep only SPES Mestre rows.

    The FGI site changes layout periodically, so the crawler deliberately uses
    document links and nearby card text rather than CSS classes tied to one theme.
    """

    own_session = session is None
    session = session or requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "SPES-Configuratore-Contabile/5.3.5 "
                "(+https://www.spesginnasticamestre.it)"
            )
        }
    )

    existing = {item.result_id: item for item in load_results()}
    warnings: list[str] = []
    scanned = 0
    matching = 0
    added = 0
    seen_document_urls: set[str] = set()

    try:
        for page_url, html in _classification_pages(session):
            if progress:
                progress(f"Analisi pagina FGI: {page_url}")
            for document in _document_links(page_url, html):
                if document["url"] in seen_document_urls:
                    continue
                seen_document_urls.add(document["url"])
                document_date = _parse_italian_date(document["context"])
                if document_date and not (start_date <= document_date <= end_date):
                    continue
                try:
                    response = session.get(document["url"], timeout=45)
                    response.raise_for_status()
                except requests.RequestException as exc:
                    warnings.append(f"Documento non scaricato: {document['url']} ({exc})")
                    continue

                scanned += 1
                if progress:
                    progress(f"Lettura classifica {scanned}: {document['name']}")
                try:
                    rows = _extract_document_rows(
                        response.content,
                        document["url"],
                        response.headers.get("content-type", ""),
                    )
                except Exception as exc:  # robust against malformed federation files
                    warnings.append(f"Documento non leggibile: {document['name']} ({exc})")
                    continue

                metadata = _metadata_from_context(document["context"], document_date)
                for row in rows:
                    if not _is_spes_row(row):
                        continue
                    matching += 1
                    result = _result_from_row(
                        row=row,
                        metadata=metadata,
                        source_url=document["url"],
                        document_name=document["name"],
                        start_date=start_date,
                    )
                    if result.result_id not in existing:
                        existing[result.result_id] = result
                        added += 1
    finally:
        if own_session:
            session.close()

    ordered = sorted(
        existing.values(),
        key=lambda item: (_date_sort_key(item.date), item.competition, item.athlete),
        reverse=True,
    )
    save_results(ordered)
    return UpdateReport(
        scanned_documents=scanned,
        matching_rows=matching,
        added_results=added,
        total_results=len(ordered),
        warnings=tuple(warnings[:50]),
    )


def export_results_csv(path: str | Path, rows: Iterable[FgiResult]) -> None:
    headers = [
        "STAGIONE",
        "DATA",
        "DISCIPLINA",
        "GARA",
        "CATEGORIA",
        "ATLETA",
        "SOCIETA",
        "POSIZIONE",
        "PUNTEGGIO",
        "ATTREZZO",
        "FONTE",
    ]
    with Path(path).open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle, delimiter=";", lineterminator="\n")
        writer.writerow(headers)
        for item in rows:
            writer.writerow(
                [
                    item.season,
                    item.date,
                    item.discipline,
                    item.competition,
                    item.category,
                    item.athlete,
                    item.club,
                    item.position,
                    item.score,
                    item.apparatus,
                    item.source_url,
                ]
            )


def export_results_xlsx(path: str | Path, rows: Iterable[FgiResult]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Risultati FGI"
    headers = [
        "STAGIONE",
        "DATA",
        "DISCIPLINA",
        "GARA",
        "CATEGORIA",
        "ATLETA",
        "SOCIETA",
        "POSIZIONE",
        "PUNTEGGIO",
        "ATTREZZO",
        "FONTE",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="073B84")
    for item in rows:
        sheet.append(
            [
                item.season,
                item.date,
                item.discipline,
                item.competition,
                item.category,
                item.athlete,
                item.club,
                item.position,
                item.score,
                item.apparatus,
                item.source_url,
            ]
        )
    widths = [13, 12, 24, 42, 24, 28, 28, 12, 14, 18, 55]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index) if index <= 26 else "A"].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def _classification_pages(session: requests.Session):
    consecutive_empty = 0
    for offset in range(0, 2000, 20):
        url = FGI_CLASSIFICATIONS_URL if offset == 0 else f"{FGI_CLASSIFICATIONS_URL}?start={offset}"
        response = session.get(url, timeout=45)
        response.raise_for_status()
        html = response.text
        links = list(_document_links(url, html))
        if not links:
            consecutive_empty += 1
        else:
            consecutive_empty = 0
        yield url, html
        if offset > 0 and consecutive_empty >= 3:
            break


def _document_links(page_url: str, html: str):
    soup = BeautifulSoup(html, "html.parser")
    seen: set[str] = set()
    for anchor in soup.find_all("a", href=True):
        href = str(anchor.get("href", "")).strip()
        label = " ".join(anchor.get_text(" ", strip=True).split())
        lowered = href.lower()
        is_download = (
            any(lowered.split("?")[0].endswith(ext) for ext in (".pdf", ".xls", ".xlsx", ".csv"))
            or "scarica" in label.lower()
            or "download" in lowered
        )
        if not is_download:
            continue
        url = urljoin(page_url, href)
        if url in seen:
            continue
        seen.add(url)
        container = anchor
        for _ in range(5):
            if container.parent is None:
                break
            container = container.parent
            text = " ".join(container.get_text(" ", strip=True).split())
            if len(text) >= 60:
                break
        context = " ".join(container.get_text(" ", strip=True).split())
        name = Path(urlparse(url).path).name or label or "classifica"
        yield {"url": url, "name": name, "context": context}


def _extract_document_rows(content: bytes, url: str, content_type: str) -> list[list[str]]:
    suffix = Path(urlparse(url).path).suffix.lower()
    content_type = content_type.lower()
    if suffix == ".pdf" or "pdf" in content_type:
        return _extract_pdf_rows(content)
    if suffix == ".xlsx" or "spreadsheetml" in content_type:
        return _extract_xlsx_rows(content)
    if suffix == ".xls" or "ms-excel" in content_type:
        return _extract_xls_rows(content)
    if suffix == ".csv" or "csv" in content_type or "text/plain" in content_type:
        return _extract_csv_rows(content)
    if content.startswith(b"%PDF"):
        return _extract_pdf_rows(content)
    return _extract_csv_rows(content)


def _extract_pdf_rows(content: bytes) -> list[list[str]]:
    results: list[list[str]] = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for row in table or []:
                    cleaned = [_clean_cell(cell) for cell in row or []]
                    if any(cleaned):
                        results.append(cleaned)
            text = page.extract_text() or ""
            lines = [" ".join(line.split()) for line in text.splitlines() if line.strip()]
            for index, line in enumerate(lines):
                if _is_spes_row([line]):
                    window = lines[max(0, index - 1) : min(len(lines), index + 2)]
                    results.append(window)
    return _deduplicate_rows(results)


def _extract_xlsx_rows(content: bytes) -> list[list[str]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    results: list[list[str]] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cleaned = [_clean_cell(value) for value in row]
            if any(cleaned):
                results.append(cleaned)
    return results


def _extract_xls_rows(content: bytes) -> list[list[str]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - dependency checked in build
        raise RuntimeError("Per leggere XLS installare xlrd") from exc
    book = xlrd.open_workbook(file_contents=content)
    results: list[list[str]] = []
    for sheet in book.sheets():
        for row_index in range(sheet.nrows):
            cleaned = [_clean_cell(value) for value in sheet.row_values(row_index)]
            if any(cleaned):
                results.append(cleaned)
    return results


def _extract_csv_rows(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    return [[_clean_cell(cell) for cell in row] for row in csv.reader(io.StringIO(text), dialect)]


def _is_spes_row(row: Iterable[str]) -> bool:
    normalized = _normalize(" ".join(str(cell) for cell in row))
    if FGI_CLUB_CODE in normalized:
        return True
    return any(_normalize(name) in normalized for name in FGI_CLUB_NAMES)


def _result_from_row(
    *,
    row: list[str],
    metadata: dict[str, str],
    source_url: str,
    document_name: str,
    start_date: date,
) -> FgiResult:
    cleaned = [cell for cell in (_clean_cell(value) for value in row) if cell]
    raw = " | ".join(cleaned)
    club_index = next(
        (
            index
            for index, cell in enumerate(cleaned)
            if FGI_CLUB_CODE in _normalize(cell)
            or any(_normalize(name) in _normalize(cell) for name in FGI_CLUB_NAMES)
        ),
        -1,
    )
    athlete = _guess_athlete(cleaned, club_index)
    position = _guess_position(cleaned)
    score = _guess_score(cleaned)
    category = metadata.get("category", "") or _guess_category(cleaned)
    apparatus = _guess_apparatus(cleaned)
    season = f"{start_date.year}/{start_date.year + 1}"
    result_date = metadata.get("date", "")
    identity = "||".join(
        [source_url, result_date, athlete, raw, FGI_CLUB_CODE]
    )
    result_id = hashlib.sha1(identity.encode("utf-8", errors="ignore")).hexdigest()
    return FgiResult(
        result_id=result_id,
        season=season,
        date=result_date,
        discipline=metadata.get("discipline", ""),
        competition=metadata.get("competition", document_name),
        category=category,
        athlete=athlete,
        club="SPES Mestre Ginnastica A.S.D. (000112)",
        position=position,
        score=score,
        apparatus=apparatus,
        source_url=source_url,
        document_name=document_name,
        raw_row=raw,
        updated_at=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    )


def _metadata_from_context(context: str, document_date: date | None) -> dict[str, str]:
    normalized = " ".join(context.split())
    discipline = next((item for item in _DISCIPLINES if item.lower() in normalized.lower()), "")
    date_text = document_date.strftime("%d/%m/%Y") if document_date else ""
    competition = normalized
    competition = re.sub(r"\bScarica\b", "", competition, flags=re.IGNORECASE).strip()
    if len(competition) > 180:
        competition = competition[:177].rstrip() + "..."
    category_match = re.search(
        r"\b(Alliev[ie]|Junior\s*\d*|Senior\s*\d*|A\d+|J\d+|S\d+|LC\d*|LD\d*|LB\d*|LA\d*)\b",
        normalized,
        flags=re.IGNORECASE,
    )
    return {
        "date": date_text,
        "discipline": discipline,
        "competition": competition,
        "category": category_match.group(0) if category_match else "",
    }


def _parse_italian_date(text: str) -> date | None:
    match = re.search(r"\b(\d{1,2})\s+([A-Za-zÀ-ÿ]+)\s+(20\d{2})\b", text)
    if match:
        day, month_name, year = match.groups()
        month = _MONTHS.get(_normalize(month_name).lower())
        if month:
            try:
                return date(int(year), month, int(day))
            except ValueError:
                pass
    match = re.search(r"\b(\d{1,2})[./-](\d{1,2})[./-](20\d{2})\b", text)
    if match:
        try:
            return date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
        except ValueError:
            return None
    return None


def _guess_athlete(cells: list[str], club_index: int) -> str:
    candidates = cells[:club_index] if club_index > 0 else cells
    for cell in reversed(candidates):
        text = re.sub(r"\b\d{2}/\d{2}/\d{4}\b", "", cell).strip(" -|")
        if _looks_like_person(text):
            return text
    for cell in cells:
        if _looks_like_person(cell) and not _is_spes_row([cell]):
            return cell
    return "Da verificare"


def _looks_like_person(value: str) -> bool:
    value = " ".join(value.split())
    if len(value) < 5 or len(value) > 80:
        return False
    if any(char.isdigit() for char in value):
        return False
    words = value.split()
    if len(words) < 2 or len(words) > 6:
        return False
    forbidden = ("classifica", "categoria", "societa", "società", "punteggio", "gara")
    lowered = value.lower()
    return not any(word in lowered for word in forbidden)


def _guess_position(cells: list[str]) -> str:
    for cell in cells[:3]:
        match = re.fullmatch(r"\s*(\d{1,3})(?:[°ª.]|\s*)\s*", cell)
        if match:
            return match.group(1)
    return ""


def _guess_score(cells: list[str]) -> str:
    numbers: list[str] = []
    for cell in cells:
        for match in re.findall(r"(?<!\d)(\d{1,3}[.,]\d{1,3})(?!\d)", cell):
            numbers.append(match)
    return numbers[-1] if numbers else ""


def _guess_category(cells: list[str]) -> str:
    text = " ".join(cells)
    match = re.search(
        r"\b(Alliev[ie]|Junior\s*\d*|Senior\s*\d*|A\d+|J\d+|S\d+|LC\d*|LD\d*|LB\d*|LA\d*)\b",
        text,
        flags=re.IGNORECASE,
    )
    return match.group(0) if match else ""


def _guess_apparatus(cells: list[str]) -> str:
    text = _normalize(" ".join(cells))
    for apparatus in (
        "corpo libero",
        "cavallo con maniglie",
        "anelli",
        "volteggio",
        "parallele",
        "sbarra",
        "trave",
        "clavette",
        "nastro",
        "cerchio",
        "palla",
        "fune",
    ):
        if _normalize(apparatus) in text:
            return apparatus.title()
    return ""


def _clean_cell(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\x00", " ").split())


def _normalize(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(value))
    return "".join(char for char in normalized if not unicodedata.combining(char)).upper()


def _deduplicate_rows(rows: list[list[str]]) -> list[list[str]]:
    seen: set[str] = set()
    result: list[list[str]] = []
    for row in rows:
        key = "|".join(row)
        if key not in seen:
            seen.add(key)
            result.append(row)
    return result


def _date_sort_key(value: str) -> str:
    try:
        return datetime.strptime(value, "%d/%m/%Y").strftime("%Y%m%d")
    except ValueError:
        return "00000000"
